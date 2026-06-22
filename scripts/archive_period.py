"""Per-period archive pipeline for ScreenAI Lab.

Produces a self-contained, human-readable archive of ONE completed
recruitment period so the lab no longer has to read the raw relational DB
or download candidate files one-by-one from the web UI.

What it writes (idempotent — re-running for the same period refreshes cleanly)::

    archive/{period_id}_{period_name_slug}/
    ├── report.xlsx                       # Candidates sheet + a per-division pair
    │                                     # (Dimension Score {Div} + Justification {Div})
    │                                     # for every division with >=1 application
    ├── documents/
    │   └── {NIM}_{FullNameSlug}/         # one folder per candidate/application
    │       ├── CV.pdf
    │       ├── Motivation_Letter.pdf
    │       ├── KHS.pdf
    │       ├── KTM.pdf|.jpg|.png
    │       ├── SWOT.pdf
    │       └── Supporting_Documents.pdf
    ├── full_db_snapshot_{YYYYMMDD_HHMMSS}.dump   # pg_dump -Fc of the WHOLE DB
    └── manifest.json

This is a standalone process — NOT part of the FastAPI app. It is read-only
against the application data: it deletes nothing and changes no schema. The
``pg_dump`` snapshot is a full-fidelity safety net of the ENTIRE database
(all periods), taken at run time — it is deliberately NOT period-scoped.

Run it from the repo root, with the same environment / ``.env`` the backend
uses (so ``DATABASE_URL`` and the ``POSTGRES_*`` vars resolve)::

    python -m scripts.archive_period --list-periods
    python -m scripts.archive_period --period-id 1
    python -m scripts.archive_period --period-id 1 --output-dir ./archive
    python -m scripts.archive_period --period-id 1 --no-dump

The DB dump shells out, by default, to the Postgres container described in
``docker-compose.yml`` (service ``db``)::

    docker compose exec -T db pg_dump -Fc -U $POSTGRES_USER $POSTGRES_DB

``--dump-mode local`` instead runs the host's own ``pg_dump``/``pg_restore``
against the server in standard libpq env vars (``PGHOST``/``PGPORT``/
``PGPASSWORD``) — useful for non-docker deployments and for testing.

Exit codes::

    0  success (including --no-dump, which marks the snapshot skipped)
    1  setup/usage error — nothing was produced (bad --period-id, etc.)
    2  archive WAS produced, but the DB snapshot failed or did not validate
       (the report + documents are complete; see manifest.json, then re-run
       after fixing Postgres access)
"""

from __future__ import annotations

import argparse
import enum
import hashlib
import json
import logging
import os
import platform
import re
import shlex
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

# Allow both `python -m scripts.archive_period` and
# `python scripts/archive_period.py` by putting the repo root on sys.path.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# Importing the models package registers every ORM mapper (User has
# back_populates to several other models), so configure-on-first-query works.
from sqlalchemy import func  # noqa: E402

import backend.models  # noqa: F401,E402  (side-effect: register all mappers)
from backend.config import settings  # noqa: E402
from backend.database import SessionLocal  # noqa: E402
from backend.models.application import Application, ApplicationStatus, Division  # noqa: E402
from backend.models.candidate import Candidate  # noqa: E402
from backend.models.document import Document, DocumentType  # noqa: E402
from backend.models.period import RecruitmentPeriod  # noqa: E402
from backend.models.rubric import Rubric  # noqa: E402
from backend.models.user import User  # noqa: E402

try:
    from dotenv import load_dotenv

    # Populate os.environ from the project .env so POSTGRES_* (needed by the
    # dump step) resolve exactly as they do for the backend. Existing env wins.
    load_dotenv(_REPO_ROOT / ".env")
except Exception:  # dotenv is a project dep, but never let its absence break a run
    pass

log = logging.getLogger("archive_period")

SCRIPT_VERSION = "1.0.0"

# Human-readable filenames for the per-candidate document folder. Keys are the
# Phase-1 DocumentType enum members; the on-disk extension is probed separately.
DOC_TYPE_LABEL: dict[DocumentType, str] = {
    DocumentType.CV: "CV",
    DocumentType.MOTIVATION_LETTER: "Motivation_Letter",
    DocumentType.KHS: "KHS",
    DocumentType.KTM: "KTM",
    DocumentType.SWOT: "SWOT",
    DocumentType.SUPPORTING_DOCS: "Supporting_Documents",
}

# Extensions a candidate file may have on disk (KTM allows images; everything
# else is pdf). See backend/utils/file_storage.py.
_DOC_EXTS = ("pdf", "jpg", "png")


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )


def slugify(value: str | None, fallback: str = "") -> str:
    """Make a filesystem-safe slug: drop path-illegal chars, spaces -> '_'."""
    s = (value or "").strip()
    # Keep unicode word chars, dot and hyphen; everything else (including the
    # Windows/Unix-illegal set <>:"/\|?* and control chars) becomes a space.
    s = re.sub(r"[^\w.\-]+", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"_+", "_", s).strip("._")
    return s or fallback


def sha256_of(path: Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for block in iter(lambda: fh.read(chunk), b""):
            h.update(block)
    return h.hexdigest()


def fmt_dt(value) -> str:
    """Render a datetime for the report; blank for None (never the str 'None')."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def cell(value):
    """Null-safe cell value: None -> '', Enum -> .value, datetime -> string.

    Division / ApplicationStatus subclass ``str``, so the ``enum.Enum`` check
    must come first — otherwise openpyxl would write ``Division.BIG_DATA``
    instead of the canonical ``big_data``.
    """
    if value is None:
        return ""
    if isinstance(value, enum.Enum):
        return value.value
    if isinstance(value, datetime):
        return fmt_dt(value)
    return value


def doc_file_on_disk(application_id: int, doc_type: DocumentType) -> Path | None:
    """Locate the physical upload by the deterministic path scheme.

    ``{upload_dir}/{application_id}/{doc_type.value}.{ext}`` — Document.file_name
    is the *original* upload name and must NOT be used to find the file. KTM may
    be jpg/png, so probe each allowed extension.
    """
    base = Path(settings.upload_dir).resolve() / str(application_id)
    for ext in _DOC_EXTS:
        candidate = base / f"{doc_type.value}.{ext}"
        if candidate.is_file():
            return candidate
    return None


# ---------------------------------------------------------------------------
# Data gathering (pure reads)
# ---------------------------------------------------------------------------

class CandidateRecord:
    """Everything the archive needs about one application, gathered once."""

    def __init__(self, app: Application, user: User, candidate: Candidate | None):
        self.app = app
        self.user = user
        self.candidate = candidate
        # Folder name: {NIM}_{FullNameSlug}; fall back to app id when NIM missing
        # so two no-NIM candidates can't collide.
        nim_part = slugify(user.nim) or f"app{app.id}"
        name_part = slugify(user.full_name) or "candidate"
        self.folder_name = f"{nim_part}_{name_part}"

    @property
    def dimension_scores(self):
        if self.candidate is None:
            return []
        return sorted(
            self.candidate.dimension_scores,
            key=lambda ds: (ds.weighted_score if ds.weighted_score is not None else -1),
            reverse=True,
        )


def latest_candidate_for(db, user_id: int) -> Candidate | None:
    """The Candidate joined to an application by user_id (NOT a FK), latest first.

    May be None — a candidate that was never evaluated still gets a report row.
    """
    return (
        db.query(Candidate)
        .filter(Candidate.user_id == user_id)
        .order_by(Candidate.created_at.desc())
        .first()
    )


def gather_records(db, period: RecruitmentPeriod) -> list[CandidateRecord]:
    """One CandidateRecord per non-draft application stamped to this period."""
    apps = (
        db.query(Application)
        .filter(
            Application.period_id == period.id,
            Application.status != ApplicationStatus.DRAFT,
        )
        .all()
    )
    records: list[CandidateRecord] = []
    for app in apps:
        user = db.query(User).filter(User.id == app.user_id).first()
        if user is None:
            # Defensive: application with no user row. Skip with a warning
            # rather than crash; nothing meaningful to archive for it.
            log.warning("application %s has no user row (user_id=%s); skipping",
                        app.id, app.user_id)
            continue
        candidate = latest_candidate_for(db, app.user_id)
        records.append(CandidateRecord(app, user, candidate))

    # Sort: evaluated candidates first (highest composite score), then the
    # unevaluated ones, tie-broken by name — so the report reads top-down.
    def sort_key(r: CandidateRecord):
        score = r.candidate.composite_score if r.candidate else None
        return (score is None, -(score or 0.0), (r.user.full_name or "").lower())

    records.sort(key=sort_key)
    return records


# ---------------------------------------------------------------------------
# report.xlsx
# ---------------------------------------------------------------------------

CANDIDATE_HEADERS = [
    "No", "Full Name", "NIM", "Email", "WhatsApp", "Faculty", "Major", "Year",
    "IPK", "Division", "Status", "Submitted At", "Composite Score",
    "Language Score",
]

# Human-readable division labels used in the per-division sheet tab names.
DIVISION_LABEL: dict[Division, str] = {
    Division.BIG_DATA: "Big Data",
    Division.CYBER_SECURITY: "Cyber Security",
    Division.GAME_TECH: "Game Tech",
    Division.GIS: "GIS",
}

# Dimension-matrix cells show the RAW per-dimension score (0-100). The weighted
# score (score x dimension weight) lives in the Justification sheet, so the matrix
# stays a clean apples-to-apples view across dimensions. Flip this to True to put
# weighted scores in the matrix instead (see the archive-pipeline report note).
MATRIX_USE_WEIGHTED = False


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    from openpyxl.styles import Font

    ws.append(headers)
    for r in rows:
        ws.append(r)

    bold = Font(bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = bold
    ws.freeze_panes = "A2"
    if rows:
        from openpyxl.utils import get_column_letter

        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # Approximate column widths from header + cell content length (capped).
    from openpyxl.utils import get_column_letter

    for col_idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for r in rows:
            val = r[col_idx - 1]
            width = max(width, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 8), 60)


def _sheet_title(prefix: str, division_label: str, used: set[str]) -> str:
    """Build an Excel sheet title, failing loudly instead of silently truncating.

    Excel caps sheet names at 31 chars and requires them unique within a
    workbook. A future, longer division label must blow up here rather than be
    quietly cut (which would also risk a name collision).
    """
    title = f"{prefix} {division_label}"
    if len(title) > 31:
        raise ValueError(
            f"Excel sheet name exceeds 31 chars and would be truncated: "
            f"{title!r} ({len(title)} chars). Shorten the division label in "
            f"DIVISION_LABEL."
        )
    if title in used:
        raise ValueError(f"duplicate Excel sheet name: {title!r}")
    used.add(title)
    return title


def _write_dim_matrix_sheet(ws, dim_names: list[str], rows: list[list]) -> None:
    """Wide per-division matrix sheet.

    Columns: Rank | Full Name | NIM | IPK | <one per dimension> | Composite | Status.
    Freezes Rank/Full Name/NIM so candidate identity stays visible while the
    dimension columns scroll, and wraps the (often long) dimension headers.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    headers = ["Rank", "Full Name", "NIM", "IPK"] + dim_names + ["Composite", "Status"]
    ws.append(headers)
    for r in rows:
        ws.append(r)

    n_fixed_left = 4  # Rank, Full Name, NIM, IPK
    n_dims = len(dim_names)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold
        if n_fixed_left < col_idx <= n_fixed_left + n_dims:
            # Wrap long dimension headers so the matrix stays narrow.
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.row_dimensions[1].height = 48

    # Freeze the column AFTER NIM (col D) and row 2: keeps Rank/Full Name/NIM and
    # the header row pinned while the dimension columns scroll.
    ws.freeze_panes = "D2"
    if rows:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    for col_idx in range(1, len(headers) + 1):
        if col_idx == 1:  # Rank
            width = 6
        elif n_fixed_left < col_idx <= n_fixed_left + n_dims:  # dimension cols
            width = 14  # fixed-narrow; the wrapped header carries the long name
        else:  # Full Name / NIM / IPK / Composite / Status — size to content
            width = len(str(headers[col_idx - 1]))
            for r in rows:
                v = r[col_idx - 1]
                width = max(width, len(str(v)) if v not in (None, "") else 0)
            width = min(max(width + 2, 8), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _justification_cell(ds) -> str:
    """Wide-justification cell text for one DimensionScore.

    The candidate's ``justification`` for that dimension (blank, never ``"None"``,
    when null), prefixed ``[OVERRIDE] `` when the score was manually overridden.
    An overridden dimension with no justification renders as just ``[OVERRIDE]``.
    """
    text = "" if ds.justification is None else str(ds.justification)
    if ds.is_override:
        return f"[OVERRIDE] {text}" if text else "[OVERRIDE]"
    return text


def _write_justification_matrix_sheet(ws, dim_names: list[str], rows: list[list]) -> None:
    """Wide per-division justification sheet, aligned with the Dimension Score one.

    Columns: Rank | Full Name | NIM | <one per dimension>. Each dimension cell is
    that candidate's wrapped justification text for that dimension (see
    ``_justification_cell``). Rows are passed in already matching the paired
    ``Dimension Score`` sheet row-for-row. Freezes Rank/Full Name/NIM, wraps and
    top-aligns the justification cells, and leaves row heights unset so Excel can
    grow them to fit the wrapped paragraphs.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    headers = ["Rank", "Full Name", "NIM"] + dim_names
    ws.append(headers)
    for r in rows:
        ws.append(r)

    n_fixed_left = 3  # Rank, Full Name, NIM
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold
        if col_idx > n_fixed_left:
            # Wrap long dimension headers so the columns stay readable.
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.row_dimensions[1].height = 48

    # Freeze the first dimension column (col D) and row 2: keeps Rank/Full Name/NIM
    # and the header row pinned while the justification columns scroll.
    ws.freeze_panes = "D2"
    if rows:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # Wrap + top-align every justification cell. No fixed data-row height is set,
    # so Excel grows each row to fit its longest wrapped justification.
    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(n_fixed_left + 1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top")

    for col_idx in range(1, len(headers) + 1):
        if col_idx == 1:  # Rank
            width = 6
        elif col_idx <= n_fixed_left:  # Full Name / NIM — size to content
            width = len(str(headers[col_idx - 1]))
            for r in rows:
                v = r[col_idx - 1]
                width = max(width, len(str(v)) if v not in (None, "") else 0)
            width = min(max(width + 2, 8), 40)
        else:  # justification (dimension) columns — generous, fixed width
            width = 60
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _division_dimension_columns(div_records: list[CandidateRecord], db):
    """Resolve the ordered dimension columns for one division.

    Order comes from the rubric (``Rubric.dimensions`` by ``dimension.id``),
    resolved via the division's candidates' ``rubric_id`` — NOT the order rows
    happen to appear in ``DimensionScore`` — so columns are stable across runs.
    If more than one rubric shows up in a division the columns are the union,
    ordered by ``(rubric_id, dimension_id)`` (and a warning is logged).

    Returns ``(ordered_dimension_ids, dimension_names, rubric_ids)``.
    """
    col_meta: dict[int, tuple[int, str]] = {}  # dimension_id -> (rubric_id, name)
    rubric_ids = sorted({
        r.candidate.rubric_id for r in div_records
        if r.candidate is not None and r.candidate.rubric_id is not None
    })
    for rid in rubric_ids:
        rubric = db.get(Rubric, rid)
        if rubric is None:
            continue
        for dim in rubric.dimensions:
            col_meta.setdefault(dim.id, (rid, dim.name))

    # Safety net: never silently drop a dimension a candidate was actually scored
    # on but which the resolved rubric(s) don't cover (deleted rubric/dimension,
    # or a legacy candidate with a null rubric_id). Such columns sort last.
    for r in div_records:
        if r.candidate is None:
            continue
        for ds in r.candidate.dimension_scores:
            if ds.dimension_id not in col_meta:
                name = (ds.dimension.name if ds.dimension is not None
                        else f"(deleted dimension #{ds.dimension_id})")
                col_meta[ds.dimension_id] = (
                    ds.rubric_id if ds.rubric_id is not None else 10 ** 9, name)

    ordered_dim_ids = sorted(col_meta, key=lambda d: (col_meta[d][0], d))
    dim_names = [col_meta[d][1] for d in ordered_dim_ids]
    return ordered_dim_ids, dim_names, rubric_ids


def build_report(path: Path, records: list[CandidateRecord], db) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    used_titles: set[str] = set()

    # --- Sheet 1: Candidates (unchanged) ---
    cand_ws = wb.active
    cand_ws.title = "Candidates"
    used_titles.add(cand_ws.title)
    cand_rows: list[list] = []
    for i, r in enumerate(records, start=1):
        u, c = r.user, r.candidate
        cand_rows.append([
            i,
            cell(u.full_name),
            cell(u.nim),
            cell(u.email),
            cell(u.whatsapp),
            cell(u.faculty),
            cell(u.major),
            cell(u.year),
            cell(u.ipk),
            cell(r.app.division),
            cell(r.app.status),
            fmt_dt(r.app.submitted_at),
            cell(c.composite_score) if c else "",
            cell(c.language_score) if c else "",
        ])
    _write_sheet(cand_ws, CANDIDATE_HEADERS, cand_rows)

    # --- Per-division sheet pairs, in Division enum order ---
    pair_count = 0
    for div in Division:
        div_records = [r for r in records if r.app.division == div]
        if not div_records:
            continue  # only divisions with >=1 application get a sheet pair
        label = DIVISION_LABEL[div]

        ordered_dim_ids, dim_names, rubric_ids = _division_dimension_columns(
            div_records, db)
        if len(rubric_ids) > 1:
            log.warning("division %s spans %d rubrics %s; dimension columns are "
                        "the union ordered by (rubric_id, dimension_id)",
                        div.value, len(rubric_ids), rubric_ids)

        # Scored = has a Candidate row AND a non-null composite. Scored rows are
        # ranked by composite descending; unscored rows go last with blank
        # Rank/dimension/Composite cells (Status shows *why* they're unscored).
        scored, unscored = [], []
        for r in div_records:
            if r.candidate is not None and r.candidate.composite_score is not None:
                scored.append(r)
            else:
                unscored.append(r)
        scored.sort(key=lambda r: r.candidate.composite_score, reverse=True)

        # Build both per-division sheets from the SAME scored/unscored ordering so
        # the score matrix and the justification matrix line up row-for-row and
        # share the rubric-ordered dimension columns.
        matrix_rows: list[list] = []
        just_rows: list[list] = []
        for rank, r in enumerate(scored, start=1):
            by_dim = {}
            by_dim_just = {}
            for ds in r.candidate.dimension_scores:
                by_dim[ds.dimension_id] = (
                    ds.weighted_score if MATRIX_USE_WEIGHTED else ds.score)
                by_dim_just[ds.dimension_id] = _justification_cell(ds)
            dim_cells = [cell(by_dim.get(did)) for did in ordered_dim_ids]
            matrix_rows.append(
                [rank, cell(r.user.full_name), cell(r.user.nim), cell(r.user.ipk)]
                + dim_cells
                + [cell(r.candidate.composite_score), cell(r.app.status)]
            )
            just_rows.append(
                [rank, cell(r.user.full_name), cell(r.user.nim)]
                + [by_dim_just.get(did, "") for did in ordered_dim_ids]
            )
        for r in unscored:
            matrix_rows.append(
                ["", cell(r.user.full_name), cell(r.user.nim), cell(r.user.ipk)]
                + ["" for _ in ordered_dim_ids]
                + ["", cell(r.app.status)]
            )
            just_rows.append(
                ["", cell(r.user.full_name), cell(r.user.nim)]
                + ["" for _ in ordered_dim_ids]
            )

        ms = wb.create_sheet(_sheet_title("Dimension Score", label, used_titles))
        _write_dim_matrix_sheet(ms, dim_names, matrix_rows)

        # Justification sheet: wide, one row per candidate, justifications as
        # dimension columns — same rows/columns as the Dimension Score sheet.
        js = wb.create_sheet(_sheet_title("Justification", label, used_titles))
        _write_justification_matrix_sheet(js, dim_names, just_rows)
        pair_count += 1

    wb.save(path)
    log.info("wrote report.xlsx (%d candidate rows, %d division sheet-pair(s))",
             len(cand_rows), pair_count)


# ---------------------------------------------------------------------------
# documents/
# ---------------------------------------------------------------------------

def copy_documents(db, work_dir: Path, records: list[CandidateRecord]) -> tuple[list[dict], int]:
    """Copy the physical upload for each present doc type into per-candidate folders.

    Returns ``(per_candidate_manifest, total_files_copied)``. A candidate folder
    is always created (even if empty) so the subfolder count == candidate count.
    Doc types whose Document row exists but whose file is absent on disk are
    recorded under ``missing`` (the run does not fail).
    """
    docs_root = work_dir / "documents"
    docs_root.mkdir(parents=True, exist_ok=True)

    out: list[dict] = []
    total = 0
    for r in records:
        cand_dir = docs_root / r.folder_name
        cand_dir.mkdir(parents=True, exist_ok=True)

        doc_rows = (
            db.query(Document)
            .filter(Document.application_id == r.app.id)
            .all()
        )
        files: list[dict] = []
        missing: list[dict] = []
        for doc in doc_rows:
            dt = doc.doc_type  # ORM maps stored enum-name back to DocumentType
            label = DOC_TYPE_LABEL.get(dt, dt.value)
            src = doc_file_on_disk(r.app.id, dt)
            if src is None:
                expected = str(
                    Path(settings.upload_dir).resolve() / str(r.app.id)
                    / f"{dt.value}.<{('|'.join(_DOC_EXTS))}>"
                )
                missing.append({
                    "doc_type": dt.value,
                    "original_file_name": doc.file_name,
                    "reason": "file not found on disk",
                    "expected_path": expected,
                })
                continue
            dest = cand_dir / f"{label}{src.suffix.lower()}"
            shutil.copy2(src, dest)
            files.append({
                "doc_type": dt.value,
                "file": f"{r.folder_name}/{dest.name}",
                "size_bytes": dest.stat().st_size,
                "sha256": sha256_of(dest),
            })
            total += 1

        out.append({
            "application_id": r.app.id,
            "user_id": r.user.id,
            "nim": r.user.nim,
            "full_name": r.user.full_name,
            "folder": r.folder_name,
            "files": files,
            "missing": missing,
        })
        if missing:
            log.warning("application %s (%s): %d expected document(s) missing on disk",
                        r.app.id, r.user.nim or "no-NIM", len(missing))

    log.info("copied %d document file(s) into %d candidate folder(s)",
             total, len(records))
    return out, total


# ---------------------------------------------------------------------------
# full-DB pg_dump snapshot
# ---------------------------------------------------------------------------

def _run(cmd: list[str], *, stdout=None, stdin=None, env=None) -> subprocess.CompletedProcess:
    return subprocess.run(
        cmd,
        stdout=stdout if stdout is not None else subprocess.PIPE,
        stderr=subprocess.PIPE,
        stdin=stdin,
        cwd=str(_REPO_ROOT),
        env=env,
    )


def _capture_pg_dump_version(dump_mode: str, compose: list[str], db_service: str,
                             env) -> str | None:
    """Best-effort `pg_dump --version` via the same mode used for the dump."""
    try:
        if dump_mode == "docker":
            cmd = compose + ["exec", "-T", db_service, "pg_dump", "--version"]
        else:
            cmd = ["pg_dump", "--version"]
        proc = _run(cmd, env=env)
        if proc.returncode == 0:
            return proc.stdout.decode(errors="replace").strip()
    except FileNotFoundError:
        return None
    return None


def _libpq_from_database_url() -> dict:
    """Parse ``settings.DATABASE_URL`` into libpq connection parts (best-effort).

    Inside the backend container DATABASE_URL already resolves to ``db:5432``
    with the right credentials, so ``--dump-mode local`` can derive everything it
    needs to reach Postgres without any extra ``PG*`` env. Returns ``{}`` for a
    non-postgres URL (e.g. the sqlite dev default), in which case the caller
    falls back to the ``POSTGRES_*`` env vars.
    """
    try:
        from sqlalchemy.engine import make_url

        url = make_url(settings.database_url)
    except Exception:
        return {}
    if not url.drivername.startswith("postgresql"):
        return {}
    return {
        "host": url.host,
        "port": str(url.port) if url.port else None,
        "user": url.username,
        "password": url.password,
        "dbname": url.database,
    }


def run_db_dump(work_dir: Path, args) -> tuple[dict, bool]:
    """Write a full-DB ``pg_dump -Fc`` snapshot and validate it.

    Returns ``(manifest_fragment, ok)``. ``ok`` is False when the dump could not
    be produced or did not validate; the caller maps that to exit code 2.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dump_name = f"full_db_snapshot_{ts}.dump"
    dump_path = work_dir / dump_name

    frag: dict = {
        "skipped": False,
        "scope": "FULL DATABASE (all periods) — NOT scoped to this period; "
                 "full-fidelity safety-net snapshot taken at run time",
        "file": dump_name,
        "format": "pg_dump custom format (-Fc)",
        "mode": args.dump_mode,
    }

    user = os.environ.get("POSTGRES_USER")
    db_name = os.environ.get("POSTGRES_DB")
    password = os.environ.get("POSTGRES_PASSWORD")

    compose = shlex.split(args.compose_cmd)
    env = os.environ.copy()

    # --dump-mode local: when the libpq PG* env vars aren't already set, derive
    # the connection from DATABASE_URL so an in-container run needs no extra env
    # (DATABASE_URL there already points at db:5432 with the right creds).
    # Explicit PG* / POSTGRES_* always win. Docker mode is unchanged: it talks to
    # the db container via `docker compose exec`, so DATABASE_URL host/port don't
    # apply there.
    if args.dump_mode == "local":
        parsed = _libpq_from_database_url()
        user = os.environ.get("PGUSER") or user or parsed.get("user")
        db_name = os.environ.get("PGDATABASE") or db_name or parsed.get("dbname")
        password = os.environ.get("PGPASSWORD") or password or parsed.get("password")
        if not env.get("PGHOST") and parsed.get("host"):
            env["PGHOST"] = parsed["host"]
        if not env.get("PGPORT") and parsed.get("port"):
            env["PGPORT"] = parsed["port"]
        if parsed.get("host"):
            frag["connection"] = (
                f"{parsed['host']}:{parsed.get('port') or '5432'}/{db_name}")

    if password:
        env["PGPASSWORD"] = password

    if not user or not db_name:
        frag["error"] = (
            "could not resolve the Postgres user/database to dump. Set "
            "POSTGRES_USER / POSTGRES_DB (or PGUSER / PGDATABASE), ensure "
            "DATABASE_URL is a postgresql:// URL for --dump-mode local, or pass "
            "--no-dump."
        )
        log.error(frag["error"])
        return frag, False

    frag["database"] = db_name
    frag["pg_dump_version"] = _capture_pg_dump_version(
        args.dump_mode, compose, args.db_service, env)

    # Build the dump + validate commands for the chosen mode.
    if args.dump_mode == "docker":
        base = compose + ["exec", "-T"]
        if password:
            base += ["-e", f"PGPASSWORD={password}"]
        base += [args.db_service]
        dump_cmd = base + ["pg_dump", "-Fc", "-U", user, db_name]
        # pg_restore --list with no file argument reads the archive from stdin.
        validate_cmd = compose + ["exec", "-T", args.db_service, "pg_restore", "--list"]
        frag["command"] = " ".join(shlex.quote(c) for c in dump_cmd) + f" > {dump_name}"
    else:  # local: host pg_dump/pg_restore using libpq PGHOST/PGPORT/PGPASSWORD
        dump_cmd = ["pg_dump", "-Fc", "-U", user, db_name]
        validate_cmd = ["pg_restore", "--list", str(dump_path)]
        frag["command"] = " ".join(shlex.quote(c) for c in dump_cmd) + f" > {dump_name}"

    log.info("running DB snapshot (%s mode): %s", args.dump_mode, frag["command"])
    try:
        with open(dump_path, "wb") as out_fh:
            proc = _run(dump_cmd, stdout=out_fh, env=env)
    except FileNotFoundError as exc:
        frag["error"] = f"could not launch dump command: {exc}"
        log.error(frag["error"])
        return frag, False

    if proc.returncode != 0:
        frag["error"] = (
            f"pg_dump exited {proc.returncode}: "
            f"{proc.stderr.decode(errors='replace').strip()[:500]}"
        )
        log.error(frag["error"])
        # Leave the (partial) file out of the way so it isn't mistaken for valid.
        if dump_path.exists():
            dump_path.unlink()
        frag["file"] = None
        return frag, False

    size = dump_path.stat().st_size
    frag["size_bytes"] = size
    frag["sha256"] = sha256_of(dump_path)

    # Validate: pg_restore --list must exit 0 and print a non-empty TOC.
    try:
        if args.dump_mode == "docker":
            with open(dump_path, "rb") as in_fh:
                vproc = _run(validate_cmd, stdin=in_fh, env=env)
        else:
            vproc = _run(validate_cmd, env=env)
    except FileNotFoundError as exc:
        frag["validated"] = False
        frag["validation_error"] = f"could not launch pg_restore: {exc}"
        log.error(frag["validation_error"])
        return frag, False

    toc = vproc.stdout.decode(errors="replace").splitlines()
    toc_entries = [ln for ln in toc if ln.strip() and not ln.lstrip().startswith(";")]
    frag["validated"] = vproc.returncode == 0 and len(toc_entries) > 0
    frag["pg_restore_toc_entry_count"] = len(toc_entries)
    if not frag["validated"]:
        frag["validation_error"] = (
            f"pg_restore --list exited {vproc.returncode} / "
            f"{len(toc_entries)} TOC entries: "
            f"{vproc.stderr.decode(errors='replace').strip()[:300]}"
        )
        log.error(frag["validation_error"])
        return frag, False

    log.info("DB snapshot OK: %s (%d bytes, %d TOC entries)",
             dump_name, size, len(toc_entries))
    return frag, True


# ---------------------------------------------------------------------------
# manifest.json
# ---------------------------------------------------------------------------

def period_meta(period: RecruitmentPeriod) -> dict:
    return {
        "id": period.id,
        "name": period.name,
        "start_date": fmt_dt(period.start_date),
        "submission_end_date": fmt_dt(period.submission_end_date),
        "evaluation_end_date": fmt_dt(period.evaluation_end_date),
        "end_date": fmt_dt(period.end_date),
        "is_active": bool(period.is_active),
        "threshold_n": period.threshold_n,
    }


def build_manifest(work_dir: Path, period: RecruitmentPeriod,
                   records: list[CandidateRecord], doc_manifest: list[dict],
                   doc_count: int, report_path: Path, dump_frag: dict) -> dict:
    missing_flat = []
    for entry in doc_manifest:
        for m in entry["missing"]:
            missing_flat.append({
                "application_id": entry["application_id"],
                "nim": entry["nim"],
                "full_name": entry["full_name"],
                **m,
            })

    manifest = {
        "script_version": SCRIPT_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "period": period_meta(period),
        "candidate_count": len(records),
        "document_count": doc_count,
        "report": {
            "file": report_path.name,
            "size_bytes": report_path.stat().st_size,
            "sha256": sha256_of(report_path),
        },
        "documents": doc_manifest,
        "missing_documents": missing_flat,
        "db_snapshot": dump_frag,
        "tools": {
            "script_version": SCRIPT_VERSION,
            "python_version": platform.python_version(),
            "pg_dump_version": dump_frag.get("pg_dump_version"),
        },
    }
    (work_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    return manifest


# ---------------------------------------------------------------------------
# --list-periods
# ---------------------------------------------------------------------------

def list_periods(db) -> int:
    periods = db.query(RecruitmentPeriod).order_by(RecruitmentPeriod.id).all()
    if not periods:
        print("No recruitment periods found.")
        return 0
    counts = dict(
        db.query(Application.period_id, func.count())
        .filter(Application.status != ApplicationStatus.DRAFT)
        .group_by(Application.period_id)
        .all()
    )
    print(f"{'ID':>4}  {'Active':<6}  {'Apps':>4}  {'Start':<19}  {'End':<19}  Name")
    print("-" * 100)
    for p in periods:
        print(f"{p.id:>4}  {('yes' if p.is_active else ''):<6}  "
              f"{counts.get(p.id, 0):>4}  {fmt_dt(p.start_date):<19}  "
              f"{fmt_dt(p.end_date):<19}  {p.name}")
    return 0


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="archive_period",
        description="Build a self-contained archive of one recruitment period.",
    )
    parser.add_argument("--period-id", type=int,
                        help="ID of the period to archive (see --list-periods).")
    parser.add_argument("--output-dir", default="./archive",
                        help="Where the {id}_{slug} archive folder is written "
                             "(default: ./archive).")
    parser.add_argument("--no-dump", action="store_true",
                        help="Skip the full-DB pg_dump snapshot.")
    parser.add_argument("--dump-mode", choices=("docker", "local"), default="docker",
                        help="How to reach Postgres for the dump: 'docker' "
                             "(docker compose exec -T db pg_dump, default) or "
                             "'local' (host pg_dump via libpq PGHOST/PGPORT/...).")
    parser.add_argument("--db-service", default="db",
                        help="docker-compose Postgres service name (default: db).")
    parser.add_argument("--compose-cmd", default="docker compose",
                        help="Base docker compose command (default: 'docker compose').")
    parser.add_argument("--list-periods", action="store_true",
                        help="Print id, name, dates and non-draft app count, then exit.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    _setup_logging()
    args = parse_args(argv)

    db = SessionLocal()
    try:
        if args.list_periods:
            return list_periods(db)

        if args.period_id is None:
            log.error("either --list-periods or --period-id is required")
            return 1

        period = (
            db.query(RecruitmentPeriod)
            .filter(RecruitmentPeriod.id == args.period_id)
            .first()
        )
        if period is None:
            log.error("no recruitment period with id=%s (try --list-periods)",
                      args.period_id)
            return 1

        records = gather_records(db, period)
        log.info("period %s '%s': %d non-draft application(s)",
                 period.id, period.name, len(records))
        if not records:
            log.warning("period has no non-draft applications; archiving anyway "
                        "(empty report + empty documents tree)")

        output_dir = Path(args.output_dir).resolve()
        output_dir.mkdir(parents=True, exist_ok=True)
        slug = f"{period.id}_{slugify(period.name, fallback='period')}"
        final_dir = output_dir / slug
        work_dir = output_dir / f".{slug}.partial"

        # Idempotent: clear any leftover partial dir, build fresh, swap at the end.
        if work_dir.exists():
            shutil.rmtree(work_dir)
        work_dir.mkdir(parents=True)

        report_path = work_dir / "report.xlsx"
        build_report(report_path, records, db)
        doc_manifest, doc_count = copy_documents(db, work_dir, records)

        dump_ok = True
        if args.no_dump:
            dump_frag = {"skipped": True, "reason": "--no-dump"}
            log.info("DB snapshot skipped (--no-dump)")
        else:
            dump_frag, dump_ok = run_db_dump(work_dir, args)

        build_manifest(work_dir, period, records, doc_manifest, doc_count,
                       report_path, dump_frag)

        # Finalize: swap the freshly built .partial dir into place without ever
        # leaving the previous archive half-replaced. Renaming a directory over a
        # NON-EMPTY target fails with ENOTEMPTY on POSIX, so move any existing
        # archive aside first, rename the new one in, then delete the old copy.
        # On any failure the previous archive is restored and left intact; after
        # success neither a .partial nor a .old-* directory remains.
        old_dir = None
        if final_dir.exists():
            old_dir = output_dir / f"{slug}.old-{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            final_dir.rename(old_dir)
        try:
            work_dir.rename(final_dir)
        except OSError:
            # Roll back: restore the previous archive if we'd moved it aside.
            if old_dir is not None and not final_dir.exists():
                old_dir.rename(final_dir)
            raise
        if old_dir is not None:
            shutil.rmtree(old_dir)

        log.info("archive ready: %s", final_dir)
        if not args.no_dump and not dump_ok:
            log.error("archive written, but the DB snapshot FAILED — see "
                      "manifest.json, then re-run after fixing Postgres access")
            return 2
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
